from django.shortcuts import render, redirect,get_object_or_404
from django.contrib.auth import authenticate, login,logout
from django.contrib import messages
from django.contrib.auth.models import User
from django.contrib.auth.decorators import user_passes_test,login_required
from django.contrib import messages
from django.views.decorators.cache import never_cache
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Sum, Count, F, DecimalField, Case, When, Value
from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
import csv
from datetime import datetime, timedelta
from productsapp.models import Product, ProductVariant
from couponsapp.models import Coupon, CouponUsage
from openpyxl import Workbook
from orders.models import Order, OrderItem
from decimal import Decimal
from weasyprint import HTML 
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from django.utils import timezone
from datetime import timedelta


def is_admin(user):
    return user.is_staff

@never_cache
def admin_login(request):
    form_data = {'username': ''}  # Initialize form data

    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        form_data['username'] = username  # Retain username in form data

        user = authenticate(request, username=username, password=password)

        if user is not None and user.is_staff:
            login(request, user)
            return redirect('admin_dashboard')
        else:
            messages.error(request, 'Invalid credentials or not an admin user.')

    return render(request, 'admin/admin_login.html', {'form_data': form_data})




@user_passes_test(is_admin)
@login_required(login_url='admin_login')
def admin_dashboard(request):
    # Default date range (last 30 days)
    end_date = timezone.now()
    start_date = end_date - timedelta(days=30)
    date_range = None

    # Handle date range filtering
    if request.method == 'POST':
        date_range = request.POST.get('date_range')
        custom_start = request.POST.get('custom_start')
        custom_end = request.POST.get('custom_end')

        if date_range == 'custom':
            try:
                start_date = timezone.datetime.strptime(custom_start, '%Y-%m-%d').date()
                end_date = timezone.datetime.strptime(custom_end, '%Y-%m-%d').date()
                
                # Validate dates
                if start_date > timezone.now().date():
                    messages.error(request, "Start date cannot be in the future")
                    print('Start date cannot be in the future')
                    return redirect('admin_dashboard')
                if end_date > timezone.now().date():
                    messages.error(request, "End date cannot be in the future")
                    print('End date cannot be in the future')
                    return redirect('admin_dashboard')
                if start_date > end_date:
                    messages.error(request, "Start date cannot be after end date")
                    print('Start date cannot be after end date')
                    return redirect('admin_dashboard')
                    
            except ValueError:
                messages.error(request, "Invalid date format")
                return redirect('admin_dashboard')
        else:
            end_date = timezone.now().date()
            if date_range == '1_day':
                start_date = end_date - timedelta(days=1)
            elif date_range == '1_week':
                start_date = end_date - timedelta(weeks=1)
            elif date_range == '1_month':
                start_date = end_date - timedelta(days=30)

    # Get COMPLETED orders within the date range
    completed_orders = Order.objects.filter(
        status='completed',
        created_at__date__range=[start_date, end_date]
    )

    # Calculate totals using actual order prices (not recalculating offers)
    total_sales = Decimal('0.00')
    total_discount = Decimal('0.00')
    
    for order in completed_orders:
        # Get only delivered/return_requested/return_denied items
        valid_items = order.items.filter(
            status__in=['delivered', 'return_requested', 'return_denied']
        )
        
        # Calculate using the actual prices stored in OrderItems
        order_total = sum(item.price * item.quantity for item in valid_items)
        total_sales += order_total
        
        # Use the actual coupon discount stored with the order
        if order.coupon:
            total_discount += order.balance_refund

    # Other dashboard data
    total_users = User.objects.filter(is_superuser=False).count()
    total_orders_count = completed_orders.count()

    order_status_counts = {
        'pending': Order.objects.filter(status='pending').count(),
        'completed': completed_orders.count(),
        'canceled': Order.objects.filter(status='canceled').count(),
        'returned': Order.objects.filter(status='returned').count(),
    }

    item_status_counts = {
        'delivered': OrderItem.objects.filter(status='delivered').count(),
        'canceled': OrderItem.objects.filter(status='canceled').count(),
        'return_requested': OrderItem.objects.filter(status='return_requested').count(),
        'returned': OrderItem.objects.filter(status='returned').count(),
    }

    # Top selling categories (only from completed orders)
    top_selling_categories = (
        OrderItem.objects.filter(
            order__status='completed',
            status__in=['delivered', 'return_requested', 'return_denied']
        )
        .values('product__category__category_name')
        .annotate(total_sold=Sum('quantity'))
        .order_by('-total_sold')[:5]
    )

    # Top selling products (only from completed orders)
    top_selling_products = Product.objects.annotate(
        total_sold=Sum(
            Case(
                When(
                    variants__orderitem__order__status='completed',
                    variants__orderitem__status__in=['delivered', 'return_requested', 'return_denied'],
                    then=F('variants__orderitem__quantity')
                ),
                default=Value(0),
                output_field=DecimalField()
            )
        )
    ).order_by('-total_sold')[:5]

    low_stock_products = ProductVariant.objects.filter(stock__lt=10)

    context = {
        'total_users': total_users,
        'total_sales': total_sales - total_discount,
        'total_orders': total_orders_count,
        'total_discount': total_discount,
        'order_status_counts': order_status_counts,
        'item_status_counts': item_status_counts,
        'top_selling_categories': top_selling_categories,
        'top_selling_products': top_selling_products,
        'low_stock_products': low_stock_products,
        'start_date': start_date.strftime('%Y-%m-%d'),
        'end_date': end_date.strftime('%Y-%m-%d'),
    }
    print(f"Start Date: {start_date}, End Date: {end_date}")
    return render(request, 'admin/dashboard.html', context)


@user_passes_test(is_admin)
@login_required(login_url='admin_login')
def generate_sales_report(request):
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    report_format = request.GET.get('format', 'pdf')

    # Validate that start_date and end_date are provided
    if not start_date_str or not end_date_str:
        return HttpResponse("Start date and end date are required", status=400)

    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        
        # Validate dates
        if start_date > end_date:
            return HttpResponse("Start date cannot be after end date", status=400)
        if end_date > timezone.now().date():
            return HttpResponse("End date cannot be in the future", status=400)
            
    except ValueError:
        return HttpResponse("Invalid date format", status=400)

    # Get completed orders in date range
    completed_orders = Order.objects.filter(
        status='completed',
        created_at__date__range=[start_date, end_date]
    )

    order_data = []
    total_sales = Decimal('0.00')
    total_discount = Decimal('0.00')

    for order in completed_orders:
        valid_items = order.items.filter(
            status__in=['delivered', 'return_requested', 'return_denied']
        )
        
        order_total = sum(item.price * item.quantity for item in valid_items)
        order_discount = order.balance_refund if order.coupon else Decimal('0.00')
        
        order_data.append({
            'id': order.id,
            'user': order.user,
            'total_before_coupon': order_total,
            'total_price': order_total  - order_discount,
            'discounted_amount': order_discount,
            'created_at': order.created_at
        })
        
        total_sales += order_total
        total_discount += order_discount

    context = {
        'order_data': order_data,
        'total_sales': total_sales - total_discount,
        'total_discount_amount': total_discount,
        'total_orders': len(order_data),
        'start_date': start_date.strftime('%Y-%m-%d'),
        'end_date': end_date.strftime('%Y-%m-%d'),
        'report_generated_at': timezone.now().strftime('%Y-%m-%d %H:%M:%S'),
    }

    if report_format == 'pdf':
        template = get_template('admin/sales_report_pdf.html')
        html = template.render(context)
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="sales_report_{start_date}_{end_date}.pdf"'
        HTML(string=html).write_pdf(response)
        return response
    elif report_format == 'excel':
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="sales_report_{start_date}_{end_date}.xlsx"'

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Sales Report'

        # Define styles
        title_font = Font(bold=True, size=16, name='Arial', color='2563EB')
        header_font = Font(bold=True, color="FFFFFF", name='Arial', size=11)
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        company_font = Font(bold=True, size=12, name='Arial', color='1e40af')
        address_font = Font(size=10, name='Arial', color='666666')
        summary_title_font = Font(bold=True, size=12, name='Arial', color='1e40af')
        summary_label_font = Font(bold=True, size=10, name='Arial', color='4b5563')
        summary_value_font = Font(bold=True, size=11, name='Arial', color='1e40af')
        normal_font = Font(name='Arial', size=10)
        amount_font = Font(name='Consolas', size=10, color='1e40af')  # Monospace font for amounts

        border = Border(
            left=Side(style='thin', color='e5e7eb'),
            right=Side(style='thin', color='e5e7eb'),
            top=Side(style='thin', color='e5e7eb'),
            bottom=Side(style='thin', color='e5e7eb')
        )

        # Alignments
        center_align = Alignment(horizontal='center', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')

        # Background fills
        light_gray_fill = PatternFill(start_color="f8f9fa", end_color="f8f9fa", fill_type="solid")
        summary_item_fill = PatternFill(start_color="f8f9fa", end_color="f8f9fa", fill_type="solid")

        # Title Section
        worksheet.merge_cells('A1:F1')
        title_cell = worksheet.cell(row=1, column=1, value="SALES REPORT")
        title_cell.font = title_font
        title_cell.alignment = center_align
        
        # Company Info Section
        worksheet.merge_cells('A3:F3')
        company_cell = worksheet.cell(row=3, column=1, value="URBAN EDGE")
        company_cell.font = company_font
        company_cell.alignment = left_align

        # Company Address
        address_text = "National Highway 66 near Calicut University\nKakkanchery Chelembra PO, Dt, Thenhipalam\nKerala 673634"
        worksheet.merge_cells('A4:F4')
        address_cell = worksheet.cell(row=4, column=1, value=address_text)
        address_cell.font = address_font
        address_cell.alignment = left_align

        # Period Info (changed from right_align to left_align)
        worksheet.merge_cells('A6:F6')
        period_cell = worksheet.cell(row=6, column=1, value=f"Period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}")
        period_cell.font = normal_font
        period_cell.alignment = left_align  # Changed from right_align to left_align

        # Summary Section
        worksheet.merge_cells('A8:F8')
        summary_title = worksheet.cell(row=8, column=1, value="Summary")
        summary_title.font = summary_title_font

        # Summary Data
        summary_data = [
            ('Total Orders', len(order_data)),
            ('Total Sales', f"₹{total_sales:,.2f}"),
            ('Total Discount', f"₹{total_discount:,.2f}")
        ]

        # Create summary grid
        start_row = 9
        for idx, (label, value) in enumerate(summary_data):
            # Labels
            cell = worksheet.cell(row=start_row + idx, column=1, value=label)
            cell.font = summary_label_font
            cell.fill = summary_item_fill
            cell.border = border
            cell.alignment = left_align  # Added left alignment
            
            # Values
            cell = worksheet.cell(row=start_row + idx, column=2, value=value)
            cell.font = summary_value_font
            cell.fill = summary_item_fill
            cell.border = border
            cell.alignment = right_align  # Changed to always use left alignment


        # Order Details Section
        detail_row = start_row + 5
        worksheet.merge_cells(f'A{detail_row}:F{detail_row}')
        detail_title = worksheet.cell(row=detail_row, column=1, value="Order Details")
        detail_title.font = summary_title_font

        # Headers
        headers = ['Order ID', 'Customer', 'Amount (₹)', 'Discount (₹)', 'After Discount (₹)', 'Date']
        header_row = detail_row + 1
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=header_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_align

        # Data rows
        for idx, order in enumerate(order_data, 1):
            row = header_row + idx
            data = [
                order['id'],
                order['user'].username,
                f"₹{order['total_before_coupon']:,.2f}",
                f"₹{order['discounted_amount']:,.2f}",
                f"₹{order['total_price']:,.2f}",
                order['created_at'].strftime('%Y-%m-%d %H:%M:%S')
            ]
            
            for col, value in enumerate(data, 1):
                cell = worksheet.cell(row=row, column=col, value=value)
                cell.border = border
                cell.alignment = right_align if isinstance(value, str) and '₹' in value else center_align
                cell.font = amount_font if isinstance(value, str) and '₹' in value else normal_font
                if row % 2 == 0:
                    cell.fill = light_gray_fill

        # Set column widths
        column_widths = [25, 20, 18, 18, 18, 22]
        for i, width in enumerate(column_widths, 1):
            worksheet.column_dimensions[get_column_letter(i)].width = width

        # Generated date at bottom
        last_row = worksheet.max_row + 2
        worksheet.merge_cells(f'A{last_row}:F{last_row}')
        generated_cell = worksheet.cell(row=last_row, column=1, value=f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}")
        generated_cell.font = normal_font
        generated_cell.alignment = center_align

        workbook.save(response)
        return response
    
    return (request, 'admin_dashboard')


@user_passes_test(is_admin)
def user_manage(request):
    users = User.objects.filter(is_superuser=False).order_by('-date_joined')

    # Pagination
    page = request.GET.get('page', 1)  # Get the page number from the URL parameters
    paginator = Paginator(users, 10)  # Show 10 products per page

    try:
        users = paginator.page(page)
    except PageNotAnInteger:
        users = paginator.page(1)
    except EmptyPage:
        users = paginator.page(paginator.num_pages)


    return render(request,'admin/users.html',{'users':users})

@user_passes_test(is_admin)
def block_user(request,user_id):
    user = get_object_or_404(User, id=user_id)
    user.is_active = False
    user.save()
    messages.success(request, f'{user.username} has been blocked.')
    return redirect('users')

@user_passes_test(is_admin)
def unblock_user(request, user_id):
    user = get_object_or_404(User, id=user_id)
    user.is_active = True
    user.save()
    messages.success(request, f'{user.username} has been unblocked.')
    return redirect('users')

@login_required(login_url='admin_login')
def admin_logout(request):
    logout(request)
    return redirect('admin_login')